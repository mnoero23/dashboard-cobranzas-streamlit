from __future__ import annotations

from datetime import date
from pathlib import Path
import re
import unicodedata

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


APP_TITLE = "Dashboard de Cobranzas y Facturación para PyMEs"
APP_SUBTITLE = "Seguimiento ejecutivo de facturación, cobranza, deuda pendiente y facturas críticas."
DATA_PATH = Path("data") / "dataset_cobranzas.xlsx"

EXPECTED_COLUMNS = [
    "id_factura",
    "id_cliente",
    "cliente",
    "tipo_cliente",
    "responsable",
    "id_servicio",
    "tipo_servicio",
    "fecha_emision",
    "fecha_vencimiento",
    "fecha_pago",
    "moneda",
    "importe_facturado",
    "importe_cobrado",
    "saldo_pendiente",
    "tipo_cambio",
    "importe_equivalente_ars",
    "saldo_equivalente_ars",
    "dias_atraso",
    "estado",
    "prioridad_cobranza",
    "proxima_accion",
    "condicion_pago_dias",
    "perfil_pago",
]

REQUIRED_FOR_DASHBOARD = [
    "fecha_emision",
    "fecha_vencimiento",
    "cliente",
    "moneda",
    "estado",
    "importe_equivalente_ars",
    "saldo_equivalente_ars",
    "saldo_pendiente",
]

DATE_COLUMNS = ["fecha_emision", "fecha_vencimiento", "fecha_pago"]
NUMERIC_COLUMNS = [
    "importe_facturado",
    "importe_cobrado",
    "saldo_pendiente",
    "tipo_cambio",
    "importe_equivalente_ars",
    "saldo_equivalente_ars",
    "dias_atraso",
    "condicion_pago_dias",
]
CHART_FILTER_DEFAULTS = {"cliente": None, "estado": None, "moneda": None, "mes": None}
CHART_CONFIG = {
    "displayModeBar": False,
    "responsive": True,
}


st.set_page_config(
    page_title=APP_TITLE,
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
    :root {
        --app-bg: #f3f6fa;
        --panel: #ffffff;
        --panel-soft: #f8fafc;
        --border: #d8e0ea;
        --border-strong: #c8d2df;
        --text: #1f2933;
        --muted: #607086;
        --accent: #245b73;
        --accent-soft: #e7f0f4;
        --success: #2f6f5e;
        --warning: #9a6b36;
        --danger: #a43f3f;
        --shadow: 0 8px 22px rgba(31, 41, 51, 0.06);
    }

    .stApp {
        background: var(--app-bg);
        color: var(--text);
    }

    [data-testid="stSidebar"] {
        background: #ffffff;
        border-right: 1px solid var(--border);
    }

    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3 {
        color: var(--text);
    }

    .main .block-container {
        max-width: 1360px;
        padding-top: 1.7rem;
        padding-bottom: 2.8rem;
    }

    .hero {
        background: linear-gradient(135deg, #ffffff 0%, #f5f8fb 100%);
        border: 1px solid var(--border);
        border-radius: 8px;
        padding: 1.25rem 1.35rem 1.15rem;
        box-shadow: var(--shadow);
        margin-bottom: 1.1rem;
    }

    .hero h1 {
        margin: 0.25rem 0 0.35rem;
        color: var(--text);
        font-size: clamp(1.85rem, 2.5vw, 2.45rem);
        line-height: 1.12;
        letter-spacing: 0;
    }

    .hero p {
        margin: 0;
        color: var(--muted);
        font-size: 1.02rem;
    }

    .badge-row {
        display: flex;
        flex-wrap: wrap;
        gap: 0.5rem;
        margin-top: 0.8rem;
    }

    .badge {
        display: inline-flex;
        align-items: center;
        gap: 0.35rem;
        border: 1px solid var(--border-strong);
        border-radius: 999px;
        background: var(--accent-soft);
        color: #24475a;
        padding: 0.38rem 0.68rem;
        font-size: 0.88rem;
        font-weight: 650;
    }

    .section-title {
        margin: 1.15rem 0 0.45rem;
        color: var(--text);
        font-size: 1.08rem;
        font-weight: 750;
    }

    .chart-help {
        margin-top: -0.15rem;
        margin-bottom: 0.55rem;
        color: var(--muted);
        font-size: 0.86rem;
    }

    .active-filters {
        border: 1px solid #cbd8e6;
        border-radius: 8px;
        background: #eef5f8;
        padding: 0.85rem 1rem;
        color: #24475a;
        box-shadow: 0 3px 12px rgba(31, 41, 51, 0.04);
    }

    .active-filters strong {
        display: block;
        margin-bottom: 0.25rem;
        color: #1f3f52;
    }

    .active-filters ul {
        margin: 0.2rem 0 0;
        padding-left: 1.15rem;
    }

    [data-testid="stMetric"] {
        background: var(--panel);
        border: 1px solid var(--border);
        border-radius: 8px;
        padding: 0.95rem 1rem;
        box-shadow: var(--shadow);
        height: 124px;
        box-sizing: border-box;
    }

    [data-testid="stMetricLabel"] {
        color: var(--muted);
        font-weight: 700;
    }

    [data-testid="stMetricValue"] {
        color: var(--text);
        font-size: 1.42rem;
        font-weight: 750;
        line-height: 1.16;
    }

    div[data-testid="stPlotlyChart"] {
        background: var(--panel);
        border: 1px solid var(--border);
        border-radius: 8px;
        padding: 0.45rem;
        box-shadow: var(--shadow);
    }

    [data-testid="stDataFrame"] {
        border: 1px solid var(--border);
        border-radius: 8px;
        box-shadow: var(--shadow);
        background: var(--panel);
    }

    .stDownloadButton button,
    .stButton button {
        border-radius: 6px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def normalize_name(value: object) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-zA-Z0-9]+", "_", text.lower())
    return re.sub(r"_+", "_", text).strip("_")


def format_ars(value: float | int | None) -> str:
    value = 0 if pd.isna(value) else float(value)
    return "$ " + f"{value:,.0f}".replace(",", ".")


def format_usd(value: float | int | None) -> str:
    value = 0 if pd.isna(value) else float(value)
    return "US$ " + f"{value:,.0f}".replace(",", ".")


def format_table_money(value: float | int | None, symbol: str = "$") -> str:
    value = 0 if pd.isna(value) else float(value)
    integer_part, decimal_part = f"{value:,.2f}".split(".")
    return f"{symbol}{integer_part.replace(',', '.')},{decimal_part}"


def format_count(value: float | int | None) -> str:
    value = 0 if pd.isna(value) else float(value)
    return f"{value:,.0f}".replace(",", ".")


def month_label(value: object) -> str:
    timestamp = pd.Timestamp(value)
    return timestamp.strftime("%m/%Y")


def find_table_range(path: Path, table_name: str = "FacturasTable") -> tuple[str, str] | None:
    workbook = load_workbook(path, read_only=False, data_only=True)
    for sheet in workbook.worksheets:
        for table in sheet.tables.values():
            if normalize_name(table.name) == normalize_name(table_name):
                return sheet.title, table.ref
    return None


def read_table_from_range(path: Path, sheet_name: str, table_ref: str) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(table_ref)
    rows = sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    )
    values = list(rows)
    if not values:
        return pd.DataFrame()
    headers = [str(header).strip() if header is not None else "" for header in values[0]]
    return pd.DataFrame(values[1:], columns=headers)


@st.cache_data(show_spinner=False)
def load_facturas(path: str) -> pd.DataFrame:
    excel_path = Path(path)
    xls = pd.ExcelFile(excel_path)

    facturas_sheet = next(
        (sheet for sheet in xls.sheet_names if normalize_name(sheet) == "facturas"),
        None,
    )
    if facturas_sheet:
        df = pd.read_excel(excel_path, sheet_name=facturas_sheet)
    else:
        table_info = find_table_range(excel_path)
        if table_info is None:
            raise ValueError("No se encontró una hoja 'Facturas' ni una tabla 'FacturasTable'.")
        sheet_name, table_ref = table_info
        df = read_table_from_range(excel_path, sheet_name, table_ref)

    rename_map = {}
    expected_by_norm = {normalize_name(col): col for col in EXPECTED_COLUMNS}
    for column in df.columns:
        normalized = normalize_name(column)
        if normalized in expected_by_norm:
            rename_map[column] = expected_by_norm[normalized]
    df = df.rename(columns=rename_map)

    missing_columns = [column for column in EXPECTED_COLUMNS if column not in df.columns]
    df.attrs["missing_columns"] = missing_columns

    for column in EXPECTED_COLUMNS:
        if column not in df.columns:
            df[column] = pd.NA

    for column in DATE_COLUMNS:
        df[column] = pd.to_datetime(df[column], errors="coerce")

    for column in NUMERIC_COLUMNS:
        df[column] = pd.to_numeric(df[column], errors="coerce").fillna(0)

    for column in ["cliente", "moneda", "estado", "prioridad_cobranza", "proxima_accion"]:
        df[column] = df[column].fillna("Sin dato").astype(str).str.strip()
        df.loc[df[column].eq(""), column] = "Sin dato"

    df["mes_emision"] = df["fecha_emision"].dt.to_period("M").dt.to_timestamp()
    return df


def validate_dashboard_columns(df: pd.DataFrame) -> list[str]:
    missing_columns = set(df.attrs.get("missing_columns", []))
    return [column for column in REQUIRED_FOR_DASHBOARD if column in missing_columns]


def init_session_state() -> None:
    st.session_state.setdefault("chart_filters", CHART_FILTER_DEFAULTS.copy())
    st.session_state.setdefault("chart_version", 0)
    st.session_state.setdefault("selected_clients", [])
    st.session_state.setdefault("selected_currencies", [])
    st.session_state.setdefault("selected_states", [])


def reset_chart_filters() -> None:
    st.session_state["chart_filters"] = CHART_FILTER_DEFAULTS.copy()
    st.session_state["chart_version"] = st.session_state.get("chart_version", 0) + 1


def reset_all_filters() -> None:
    reset_chart_filters()
    for key in ["selected_clients", "selected_currencies", "selected_states"]:
        st.session_state[key] = []
    if "selected_period" in st.session_state and "default_period" in st.session_state:
        st.session_state["selected_period"] = st.session_state["default_period"]


def sync_date_filter(valid_dates: pd.Series) -> tuple[date | None, date | None]:
    if valid_dates.empty:
        return None, None
    min_date = valid_dates.min().date()
    max_date = valid_dates.max().date()
    default_period = (min_date, max_date)
    st.session_state["default_period"] = default_period

    selected_period = st.session_state.get("selected_period")
    if (
        not isinstance(selected_period, tuple)
        or len(selected_period) != 2
        or selected_period[0] < min_date
        or selected_period[1] > max_date
    ):
        st.session_state["selected_period"] = default_period

    return min_date, max_date


def segmented_multi(label: str, options: list[str], key: str) -> list[str]:
    if hasattr(st, "segmented_control"):
        value = st.segmented_control(
            label,
            options,
            selection_mode="multi",
            key=key,
            width="stretch",
        )
        return list(value or [])
    return st.multiselect(label, options, key=key)


def render_sidebar_filters(df: pd.DataFrame) -> dict[str, object]:
    with st.sidebar:
        st.header("Filtros")
        st.caption("Los filtros vacíos se interpretan como todos los valores disponibles.")

        valid_dates = df["fecha_emision"].dropna()
        min_date, max_date = sync_date_filter(valid_dates)
        if min_date and max_date:
            selected_period = st.date_input(
                "Período",
                min_value=min_date,
                max_value=max_date,
                format="DD/MM/YYYY",
                key="selected_period",
            )
            if isinstance(selected_period, tuple) and len(selected_period) == 2:
                start_date, end_date = selected_period
            else:
                start_date, end_date = min_date, max_date
        else:
            st.warning("No hay fechas de emisión válidas para filtrar por período.")
            start_date = end_date = None

        clients = sorted(df["cliente"].dropna().unique())
        currencies = sorted(df["moneda"].dropna().unique())
        states = sorted(df["estado"].dropna().unique())

        selected_clients = st.multiselect(
            "Cliente",
            clients,
            key="selected_clients",
            placeholder="Todos los clientes",
        )
        selected_currencies = segmented_multi("Moneda", currencies, "selected_currencies")
        selected_states = segmented_multi("Estado", states, "selected_states")

        chart_filters = active_chart_filters()
        if chart_filters:
            st.markdown("**Filtros activos por gráfico**")
            for label, value in chart_filters:
                st.caption(f"{label}: {value}")

        st.button(
            "Limpiar selección de gráficos",
            icon=":material/ink_eraser:",
            on_click=reset_chart_filters,
            width="stretch",
            disabled=not chart_filters,
        )
        st.button(
            "Limpiar todos los filtros",
            icon=":material/filter_alt_off:",
            on_click=reset_all_filters,
            width="stretch",
        )

    return {
        "start_date": start_date,
        "end_date": end_date,
        "clientes": selected_clients,
        "monedas": selected_currencies,
        "estados": selected_states,
    }


def active_chart_filters() -> list[tuple[str, str]]:
    filters = st.session_state.get("chart_filters", CHART_FILTER_DEFAULTS.copy())
    active: list[tuple[str, str]] = []
    if filters.get("cliente"):
        active.append(("Cliente", filters["cliente"]))
    if filters.get("moneda"):
        active.append(("Moneda", filters["moneda"]))
    if filters.get("estado"):
        active.append(("Estado", filters["estado"]))
    if filters.get("mes") is not None:
        active.append(("Período", month_label(filters["mes"])))
    return active


def apply_filters(df: pd.DataFrame, sidebar_filters: dict[str, object]) -> pd.DataFrame:
    filtered = df.copy()

    start_date = sidebar_filters.get("start_date")
    end_date = sidebar_filters.get("end_date")
    if start_date and end_date:
        filtered = filtered[
            filtered["fecha_emision"].between(
                pd.Timestamp(start_date),
                pd.Timestamp(end_date),
                inclusive="both",
            )
        ]

    clientes = sidebar_filters.get("clientes") or []
    monedas = sidebar_filters.get("monedas") or []
    estados = sidebar_filters.get("estados") or []

    if clientes:
        filtered = filtered[filtered["cliente"].isin(clientes)]
    if monedas:
        filtered = filtered[filtered["moneda"].isin(monedas)]
    if estados:
        filtered = filtered[filtered["estado"].isin(estados)]

    chart_filters = st.session_state.get("chart_filters", CHART_FILTER_DEFAULTS.copy())
    if chart_filters.get("cliente"):
        filtered = filtered[filtered["cliente"].eq(chart_filters["cliente"])]
    if chart_filters.get("moneda"):
        filtered = filtered[filtered["moneda"].eq(chart_filters["moneda"])]
    if chart_filters.get("estado"):
        filtered = filtered[filtered["estado"].eq(chart_filters["estado"])]
    if chart_filters.get("mes") is not None:
        selected_month = pd.Timestamp(chart_filters["mes"]).to_period("M").to_timestamp()
        filtered = filtered[filtered["mes_emision"].eq(selected_month)]

    return filtered


def build_kpis(df: pd.DataFrame) -> dict[str, float | int]:
    today = pd.Timestamp(date.today())
    vencidas_mask = (df["fecha_vencimiento"] < today) & (df["saldo_equivalente_ars"] > 0)

    total_facturado_ars = df["importe_equivalente_ars"].sum()
    saldo_pendiente_ars = df["saldo_equivalente_ars"].sum()

    return {
        "total_facturado_ars": total_facturado_ars,
        "total_cobrado_ars": total_facturado_ars - saldo_pendiente_ars,
        "saldo_pendiente_ars": saldo_pendiente_ars,
        "saldo_vencido_ars": df.loc[vencidas_mask, "saldo_equivalente_ars"].sum(),
        "facturas_vencidas": int(vencidas_mask.sum()),
        "saldo_pendiente_usd": df.loc[df["moneda"].str.upper().eq("USD"), "saldo_pendiente"].sum(),
    }


def style_plot(fig, *, height: int) -> object:
    fig.update_layout(
        template="plotly_white",
        height=height,
        font=dict(color="#25313f", family="Inter, Segoe UI, Arial, sans-serif"),
        margin=dict(l=12, r=14, t=16, b=18),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
        legend_title_text="",
        paper_bgcolor="white",
        plot_bgcolor="white",
        hoverlabel=dict(bgcolor="white", font_size=12),
        dragmode="select",
    )
    fig.update_xaxes(showgrid=True, gridcolor="#edf1f5", zeroline=False, title="")
    fig.update_yaxes(showgrid=False, zeroline=False, title="")
    return fig


def first_selected_point(event: object) -> dict | None:
    selection = getattr(event, "selection", None)
    if selection is None and isinstance(event, dict):
        selection = event.get("selection")
    if selection is None:
        return None

    points = getattr(selection, "points", None)
    if points is None and isinstance(selection, dict):
        points = selection.get("points")
    if not points:
        return None
    return dict(points[0])


def read_custom_value(point: dict, index: int = 0) -> object | None:
    customdata = point.get("customdata")
    if isinstance(customdata, (list, tuple)) and len(customdata) > index:
        return customdata[index]
    return customdata


def update_chart_filter(filter_name: str, value: object | None) -> None:
    if value is None or pd.isna(value):
        return

    next_filters = st.session_state.get("chart_filters", CHART_FILTER_DEFAULTS.copy()).copy()
    if filter_name == "mes":
        value = pd.Timestamp(value).to_period("M").to_timestamp()
    else:
        value = str(value)

    if next_filters.get(filter_name) == value:
        return

    next_filters[filter_name] = value
    st.session_state["chart_filters"] = next_filters
    st.rerun()


def handle_plotly_selection(event: object, filter_name: str) -> None:
    point = first_selected_point(event)
    if not point:
        return

    if filter_name == "mes":
        value = read_custom_value(point, 0) or point.get("x")
    else:
        value = read_custom_value(point, 0) or point.get("y") or point.get("x")
    update_chart_filter(filter_name, value)


def render_header() -> None:
    st.markdown(
        f"""
        <div class="hero">
            <span class="badge">Dataset simulado · Caso de portfolio</span>
            <h1>{APP_TITLE}</h1>
            <p>{APP_SUBTITLE}</p>
            <div class="badge-row">
                <span class="badge">App interactiva desarrollada con Python, Streamlit, pandas y Plotly.</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_active_filter_banner() -> None:
    active = active_chart_filters()
    if not active:
        return

    items = "".join(f"<li>{label}: <strong>{value}</strong></li>" for label, value in active)
    st.markdown(
        f"""
        <div class="active-filters">
            <strong>Filtros activos por interacción</strong>
            <ul>{items}</ul>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.button(
        "Limpiar selección de gráficos",
        icon=":material/ink_eraser:",
        on_click=reset_chart_filters,
        width="content",
    )


def render_metrics(kpis: dict[str, float | int]) -> None:
    st.markdown('<div class="section-title">KPIs principales</div>', unsafe_allow_html=True)
    first_row = st.columns(3, gap="medium")
    first_row[0].metric("Total facturado", format_ars(kpis["total_facturado_ars"]), border=True)
    first_row[1].metric("Total cobrado", format_ars(kpis["total_cobrado_ars"]), border=True)
    first_row[2].metric("Saldo pendiente", format_ars(kpis["saldo_pendiente_ars"]), border=True)

    second_row = st.columns(3, gap="medium")
    second_row[0].metric("Saldo vencido", format_ars(kpis["saldo_vencido_ars"]), border=True)
    second_row[1].metric("Facturas vencidas", format_count(kpis["facturas_vencidas"]), border=True)
    second_row[2].metric("Saldo pendiente USD", format_usd(kpis["saldo_pendiente_usd"]), border=True)


def render_monthly_chart(df: pd.DataFrame) -> None:
    st.markdown('<div class="section-title">Evolución mensual de facturación y cobranza</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="chart-help">Seleccioná un punto o usá selección rectangular para filtrar por mes.</div>',
        unsafe_allow_html=True,
    )

    monthly = (
        df.dropna(subset=["mes_emision"])
        .groupby("mes_emision", as_index=False)
        .agg(
            **{
                "Total facturado ARS Eq": ("importe_equivalente_ars", "sum"),
                "Saldo pendiente ARS Eq": ("saldo_equivalente_ars", "sum"),
            }
        )
    )
    if monthly.empty:
        st.info("No hay datos suficientes para mostrar la evolución mensual.", icon=":material/info:")
        return

    monthly["Total cobrado ARS Eq"] = (
        monthly["Total facturado ARS Eq"] - monthly["Saldo pendiente ARS Eq"]
    )
    monthly_long = monthly.melt(
        id_vars="mes_emision",
        value_vars=["Total facturado ARS Eq", "Total cobrado ARS Eq"],
        var_name="Indicador",
        value_name="Importe",
    )
    monthly_long["mes_label"] = monthly_long["mes_emision"].dt.strftime("%m/%Y")

    fig = px.line(
        monthly_long,
        x="mes_emision",
        y="Importe",
        color="Indicador",
        markers=True,
        custom_data=["mes_emision", "mes_label", "Indicador", "Importe"],
        color_discrete_sequence=["#245b73", "#2f6f5e"],
    )
    fig.update_traces(
        hovertemplate="<b>%{customdata[1]}</b><br>%{customdata[2]}: $ %{customdata[3]:,.0f}<extra></extra>"
    )
    fig.update_yaxes(tickprefix="$ ", tickformat=",.0f")
    fig.update_xaxes(tickformat="%m/%Y")

    event = st.plotly_chart(
        style_plot(fig, height=330),
        key=f"chart_monthly_{st.session_state.get('chart_version', 0)}",
        on_select="rerun",
        selection_mode=("points", "box"),
        config=CHART_CONFIG,
        width="stretch",
    )
    handle_plotly_selection(event, "mes")


def render_client_chart(df: pd.DataFrame) -> None:
    debt_by_client = (
        df.groupby("cliente", as_index=False)["saldo_equivalente_ars"]
        .sum()
        .sort_values("saldo_equivalente_ars", ascending=False)
        .head(10)
        .sort_values("saldo_equivalente_ars", ascending=True)
    )
    st.markdown('<div class="section-title">Clientes con mayor deuda pendiente</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="chart-help">Seleccioná una barra para filtrar el dashboard por cliente.</div>',
        unsafe_allow_html=True,
    )
    if debt_by_client.empty:
        st.info("No hay deuda pendiente por cliente para mostrar.", icon=":material/info:")
        return

    fig = px.bar(
        debt_by_client,
        y="cliente",
        x="saldo_equivalente_ars",
        orientation="h",
        labels={"cliente": "", "saldo_equivalente_ars": "Saldo pendiente ARS Eq"},
        custom_data=["cliente", "saldo_equivalente_ars"],
        color_discrete_sequence=["#245b73"],
    )
    fig.update_traces(
        hovertemplate="<b>%{customdata[0]}</b><br>Saldo pendiente: $ %{customdata[1]:,.0f}<extra></extra>"
    )
    fig.update_xaxes(tickprefix="$ ", tickformat=",.0f")

    event = st.plotly_chart(
        style_plot(fig, height=360),
        key=f"chart_clients_{st.session_state.get('chart_version', 0)}",
        on_select="rerun",
        selection_mode=("points", "box"),
        config=CHART_CONFIG,
        width="stretch",
    )
    handle_plotly_selection(event, "cliente")


def render_status_chart(df: pd.DataFrame) -> None:
    status_counts = (
        df.groupby("estado", as_index=False)
        .size()
        .rename(columns={"size": "cantidad_facturas"})
        .sort_values("cantidad_facturas", ascending=True)
    )
    st.markdown('<div class="section-title">Estado general de cobranzas</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="chart-help">Seleccioná una barra para filtrar por estado.</div>',
        unsafe_allow_html=True,
    )
    if status_counts.empty:
        st.info("No hay estados para mostrar.", icon=":material/info:")
        return

    fig = px.bar(
        status_counts,
        y="estado",
        x="cantidad_facturas",
        orientation="h",
        labels={"estado": "", "cantidad_facturas": "Cantidad de facturas"},
        custom_data=["estado", "cantidad_facturas"],
        color_discrete_sequence=["#607086"],
    )
    fig.update_traces(
        hovertemplate="<b>%{customdata[0]}</b><br>Facturas: %{customdata[1]:,.0f}<extra></extra>"
    )

    event = st.plotly_chart(
        style_plot(fig, height=360),
        key=f"chart_status_{st.session_state.get('chart_version', 0)}",
        on_select="rerun",
        selection_mode=("points", "box"),
        config=CHART_CONFIG,
        width="stretch",
    )
    handle_plotly_selection(event, "estado")


def render_currency_chart(df: pd.DataFrame) -> None:
    currency_debt = (
        df.groupby("moneda", as_index=False)["saldo_equivalente_ars"]
        .sum()
        .sort_values("saldo_equivalente_ars", ascending=True)
    )
    st.markdown('<div class="section-title">Exposición de deuda por moneda</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="chart-help">Seleccioná una barra para filtrar por moneda.</div>',
        unsafe_allow_html=True,
    )
    if currency_debt.empty:
        st.info("No hay deuda por moneda para mostrar.", icon=":material/info:")
        return

    fig = px.bar(
        currency_debt,
        y="moneda",
        x="saldo_equivalente_ars",
        orientation="h",
        labels={"moneda": "", "saldo_equivalente_ars": "Saldo pendiente ARS Eq"},
        custom_data=["moneda", "saldo_equivalente_ars"],
        color_discrete_sequence=["#9a6b36"],
    )
    fig.update_traces(
        hovertemplate="<b>%{customdata[0]}</b><br>Saldo pendiente ARS Eq: $ %{customdata[1]:,.0f}<extra></extra>"
    )
    fig.update_xaxes(tickprefix="$ ", tickformat=",.0f")

    event = st.plotly_chart(
        style_plot(fig, height=280),
        key=f"chart_currency_{st.session_state.get('chart_version', 0)}",
        on_select="rerun",
        selection_mode=("points", "box"),
        config=CHART_CONFIG,
        width="stretch",
    )
    handle_plotly_selection(event, "moneda")


def render_charts(df: pd.DataFrame) -> None:
    render_monthly_chart(df)
    left, right = st.columns(2, gap="medium")
    with left:
        render_client_chart(df)
    with right:
        render_status_chart(df)
    render_currency_chart(df)


def build_critical_invoices(df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "cliente",
        "id_factura",
        "fecha_vencimiento",
        "moneda",
        "saldo_pendiente",
        "saldo_equivalente_ars",
        "dias_atraso",
        "prioridad_cobranza",
        "proxima_accion",
        "estado",
    ]
    critical = df[
        (df["saldo_equivalente_ars"] > 0)
        & (~df["estado"].str.casefold().eq("cobrada"))
    ].copy()
    if critical.empty:
        return pd.DataFrame(columns=columns)
    return critical[columns].sort_values("dias_atraso", ascending=False)


def render_critical_invoices(df: pd.DataFrame) -> None:
    st.markdown('<div class="section-title">Facturas críticas para gestionar</div>', unsafe_allow_html=True)
    critical = build_critical_invoices(df)

    if critical.empty:
        st.success(
            "No hay facturas críticas para gestionar con los filtros seleccionados.",
            icon=":material/check_circle:",
        )
        return

    display_df = critical.copy()
    display_df["fecha_vencimiento"] = display_df["fecha_vencimiento"].dt.strftime("%d/%m/%Y")
    display_df["saldo_pendiente"] = display_df.apply(
        lambda row: format_table_money(row["saldo_pendiente"], "US$")
        if str(row["moneda"]).upper() == "USD"
        else format_table_money(row["saldo_pendiente"]),
        axis=1,
    )
    display_df["saldo_equivalente_ars"] = display_df["saldo_equivalente_ars"].map(
        format_table_money
    )
    display_df["dias_atraso"] = display_df["dias_atraso"].fillna(0).astype(int)

    st.dataframe(
        display_df,
        hide_index=True,
        width="stretch",
        column_config={
            "cliente": st.column_config.TextColumn("Cliente", pinned=True),
            "id_factura": st.column_config.TextColumn("Factura"),
            "fecha_vencimiento": st.column_config.TextColumn("Vencimiento"),
            "moneda": st.column_config.TextColumn("Moneda"),
            "saldo_pendiente": st.column_config.TextColumn("Saldo pendiente"),
            "saldo_equivalente_ars": st.column_config.TextColumn("Saldo ARS Eq"),
            "dias_atraso": st.column_config.NumberColumn("Días atraso", format="%d"),
            "prioridad_cobranza": st.column_config.TextColumn("Prioridad"),
            "proxima_accion": st.column_config.TextColumn("Próxima acción"),
            "estado": st.column_config.TextColumn("Estado"),
        },
    )

    csv_df = critical.copy()
    csv_df["fecha_vencimiento"] = csv_df["fecha_vencimiento"].dt.strftime("%d/%m/%Y")
    st.download_button(
        "Descargar tabla filtrada",
        data=csv_df.to_csv(index=False).encode("utf-8-sig"),
        file_name="facturas_criticas_filtradas.csv",
        mime="text/csv",
        icon=":material/download:",
    )


def build_client_summary(df: pd.DataFrame) -> pd.DataFrame:
    today = pd.Timestamp(date.today())
    work_df = df.copy()
    work_df["factura_vencida"] = (
        (work_df["fecha_vencimiento"] < today) & (work_df["saldo_equivalente_ars"] > 0)
    )
    work_df["cobrada"] = work_df["estado"].str.casefold().eq("cobrada")
    work_df["saldo_vencido_cliente_ars"] = work_df["saldo_equivalente_ars"].where(
        work_df["factura_vencida"], 0
    )

    summary = (
        work_df.groupby("cliente", as_index=False)
        .agg(
            facturas=("id_factura", "count"),
            total_facturado_ars=("importe_equivalente_ars", "sum"),
            total_cobrado_ars=("importe_cobrado", "sum"),
            saldo_pendiente_ars=("saldo_equivalente_ars", "sum"),
            saldo_vencido_ars=("saldo_vencido_cliente_ars", "sum"),
            facturas_vencidas=("factura_vencida", "sum"),
            facturas_abiertas=("cobrada", lambda values: int((~values).sum())),
            promedio_dias_atraso=("dias_atraso", "mean"),
        )
        .sort_values("saldo_pendiente_ars", ascending=False)
    )
    summary["participacion_deuda"] = 0.0
    total_debt = summary["saldo_pendiente_ars"].sum()
    if total_debt > 0:
        summary["participacion_deuda"] = summary["saldo_pendiente_ars"] / total_debt
    return summary


def render_client_profile(df: pd.DataFrame, client_name: str) -> None:
    client_df = df[df["cliente"].eq(client_name)].copy()
    if client_df.empty:
        st.info("No hay datos para el cliente seleccionado.", icon=":material/info:")
        return

    kpis = build_kpis(client_df)
    client_type = client_df["tipo_cliente"].dropna().astype(str)
    payment_profile = client_df["perfil_pago"].dropna().astype(str)
    next_actions = (
        client_df.loc[client_df["saldo_equivalente_ars"] > 0, "proxima_accion"]
        .dropna()
        .astype(str)
        .value_counts()
    )

    st.markdown(f'<div class="section-title">Perfil de {client_name}</div>', unsafe_allow_html=True)
    cols = st.columns(4, gap="medium")
    cols[0].metric("Saldo pendiente", format_ars(kpis["saldo_pendiente_ars"]), border=True)
    cols[1].metric("Saldo vencido", format_ars(kpis["saldo_vencido_ars"]), border=True)
    cols[2].metric("Facturas vencidas", format_count(kpis["facturas_vencidas"]), border=True)
    cols[3].metric("Facturas totales", format_count(len(client_df)), border=True)

    info_cols = st.columns(3, gap="medium")
    info_cols[0].caption("Tipo de cliente")
    info_cols[0].write(client_type.mode().iat[0] if not client_type.empty else "Sin dato")
    info_cols[1].caption("Perfil de pago")
    info_cols[1].write(payment_profile.mode().iat[0] if not payment_profile.empty else "Sin dato")
    info_cols[2].caption("Próxima acción dominante")
    info_cols[2].write(next_actions.index[0] if not next_actions.empty else "Sin acción pendiente")

    detail_columns = [
        "id_factura",
        "fecha_emision",
        "fecha_vencimiento",
        "moneda",
        "importe_facturado",
        "saldo_pendiente",
        "saldo_equivalente_ars",
        "dias_atraso",
        "estado",
        "prioridad_cobranza",
        "proxima_accion",
    ]
    detail = client_df[detail_columns].sort_values(
        ["saldo_equivalente_ars", "dias_atraso"], ascending=[False, False]
    )
    display_detail = detail.copy()
    for column in ["fecha_emision", "fecha_vencimiento"]:
        display_detail[column] = display_detail[column].dt.strftime("%d/%m/%Y")
    display_detail["importe_facturado"] = display_detail.apply(
        lambda row: format_table_money(
            row["importe_facturado"], "US$" if str(row["moneda"]).upper() == "USD" else "$"
        ),
        axis=1,
    )
    display_detail["saldo_pendiente"] = display_detail.apply(
        lambda row: format_table_money(
            row["saldo_pendiente"], "US$" if str(row["moneda"]).upper() == "USD" else "$"
        ),
        axis=1,
    )
    display_detail["saldo_equivalente_ars"] = display_detail["saldo_equivalente_ars"].map(
        format_table_money
    )

    st.dataframe(
        display_detail,
        hide_index=True,
        width="stretch",
        column_config={
            "id_factura": st.column_config.TextColumn("Factura", pinned=True),
            "fecha_emision": st.column_config.TextColumn("Emisión"),
            "fecha_vencimiento": st.column_config.TextColumn("Vencimiento"),
            "moneda": st.column_config.TextColumn("Moneda"),
            "importe_facturado": st.column_config.TextColumn("Facturado"),
            "saldo_pendiente": st.column_config.TextColumn("Saldo origen"),
            "saldo_equivalente_ars": st.column_config.TextColumn("Saldo ARS Eq"),
            "dias_atraso": st.column_config.NumberColumn("Días atraso", format="%d"),
            "estado": st.column_config.TextColumn("Estado"),
            "prioridad_cobranza": st.column_config.TextColumn("Prioridad"),
            "proxima_accion": st.column_config.TextColumn("Próxima acción"),
        },
    )

    st.download_button(
        "Descargar facturas del cliente",
        data=detail.to_csv(index=False).encode("utf-8-sig"),
        file_name=f"facturas_{normalize_name(client_name)}.csv",
        mime="text/csv",
        icon=":material/download:",
        width="content",
    )


def render_clients_tab(df: pd.DataFrame) -> None:
    st.markdown('<div class="section-title">Clientes</div>', unsafe_allow_html=True)
    st.caption("Vista enfocada en concentración de deuda, vencimientos y seguimiento por cliente.")

    summary = build_client_summary(df)
    if summary.empty:
        st.info("No hay clientes para mostrar con los filtros seleccionados.", icon=":material/info:")
        return

    clients_with_debt = int((summary["saldo_pendiente_ars"] > 0).sum())
    top_client = summary.iloc[0]
    metrics = st.columns(3, gap="medium")
    metrics[0].metric("Clientes con deuda", format_count(clients_with_debt), border=True)
    metrics[1].metric(
        f"Mayor saldo · {top_client['cliente']}",
        format_ars(top_client["saldo_pendiente_ars"]),
        border=True,
    )
    metrics[2].metric("Deuda vencida", format_ars(summary["saldo_vencido_ars"].sum()), border=True)

    left, right = st.columns([0.54, 0.46], gap="medium")
    with left:
        ranked = summary.head(8).sort_values("saldo_pendiente_ars", ascending=True)
        fig = px.bar(
            ranked,
            y="cliente",
            x="saldo_pendiente_ars",
            orientation="h",
            labels={
                "cliente": "",
                "saldo_pendiente_ars": "Saldo pendiente ARS Eq",
                "facturas_vencidas": "Vencidas",
            },
            custom_data=["cliente", "saldo_pendiente_ars", "facturas_vencidas"],
            color_discrete_sequence=["#245b73"],
        )
        fig.update_traces(
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "Saldo pendiente: $ %{customdata[1]:,.0f}<br>"
                "Facturas vencidas: %{customdata[2]:,.0f}<extra></extra>"
            )
        )
        fig.update_xaxes(tickprefix="$ ", tickformat=",.0f")
        fig.update_yaxes(tickfont=dict(size=11), automargin=True)
        st.plotly_chart(style_plot(fig, height=320), config=CHART_CONFIG, width="stretch")

    with right:
        table = summary.copy()
        table["participacion_deuda"] = table["participacion_deuda"].map(lambda value: f"{value:.1%}")
        table["saldo_pendiente_ars"] = table["saldo_pendiente_ars"].map(format_table_money)
        st.dataframe(
            table[["cliente", "facturas_vencidas", "saldo_pendiente_ars", "participacion_deuda"]],
            hide_index=True,
            height=320,
            width="stretch",
            column_config={
                "cliente": st.column_config.TextColumn("Cliente", width="medium"),
                "facturas_vencidas": st.column_config.NumberColumn("Vencidas", format="%d", width="small"),
                "saldo_pendiente_ars": st.column_config.TextColumn("Saldo", width="small"),
                "participacion_deuda": st.column_config.TextColumn("% deuda", width="small"),
            },
        )

    client_options = sorted(summary["cliente"].tolist(), key=str.casefold)
    selected_client = st.selectbox(
        "Analizar cliente",
        client_options,
        index=0,
        placeholder="Seleccioná un cliente",
    )
    render_client_profile(df, selected_client)


def render_project_note() -> None:
    st.markdown('<div class="section-title">Sobre este proyecto</div>', unsafe_allow_html=True)
    st.write(
        "Este dashboard fue desarrollado como caso de portfolio para mostrar cómo una PyME "
        "puede transformar una base de facturación y cobranzas en una herramienta ejecutiva "
        "de seguimiento. Permite identificar deuda pendiente, facturas vencidas, clientes "
        "críticos, exposición por moneda y acciones prioritarias de cobranza usando datos simulados."
    )

    st.markdown("### Trabajar desde otra PC")
    st.write("Con Git y Python instalados, abrí una terminal y seguí estos pasos:")
    st.code(
        "git clone <URL_DEL_REPOSITORIO>\n"
        "cd dashboard-cobranzas-streamlit\n"
        "python -m venv .venv\n"
        "# Windows: .venv\\Scripts\\activate\n"
        "# macOS/Linux: source .venv/bin/activate\n"
        "pip install -r requirements.txt\n"
        "streamlit run app.py",
        language="bash",
    )
    st.caption(
        "Si el proyecto ya está clonado en esa PC, ejecutá `git pull` antes de iniciar. "
        "El archivo de datos debe permanecer en `data/dataset_cobranzas.xlsx`."
    )


def main() -> None:
    init_session_state()
    render_header()

    if not DATA_PATH.exists():
        st.error(
            "No se encontró el archivo de datos. Ubicá el Excel en "
            "`data/dataset_cobranzas.xlsx` y volvé a ejecutar la app.",
            icon=":material/error:",
        )
        st.stop()

    try:
        df = load_facturas(str(DATA_PATH))
    except Exception as exc:
        st.error("No se pudo leer el archivo Excel.", icon=":material/error:")
        st.exception(exc)
        st.stop()

    missing_required = validate_dashboard_columns(df)
    if missing_required:
        st.error(
            "Faltan columnas necesarias para construir el dashboard: "
            + ", ".join(f"`{column}`" for column in missing_required),
            icon=":material/error:",
        )
        st.stop()

    missing_optional = [
        column for column in EXPECTED_COLUMNS if column not in df.columns or df[column].isna().all()
    ]
    if missing_optional:
        with st.expander("Columnas opcionales no encontradas o vacías", icon=":material/info:"):
            st.write(", ".join(missing_optional))

    sidebar_filters = render_sidebar_filters(df)
    filtered = apply_filters(df, sidebar_filters)

    render_active_filter_banner()

    if filtered.empty:
        st.info("No hay datos disponibles para los filtros seleccionados.", icon=":material/info:")
        st.stop()

    overview_tab, clients_tab, invoices_tab, about_tab = st.tabs(
        ["Resumen", "Clientes", "Facturas críticas", "Proyecto"]
    )

    with overview_tab:
        kpis = build_kpis(filtered)
        render_metrics(kpis)
        render_charts(filtered)

    with clients_tab:
        render_clients_tab(filtered)

    with invoices_tab:
        render_critical_invoices(filtered)

    with about_tab:
        render_project_note()


if __name__ == "__main__":
    main()
