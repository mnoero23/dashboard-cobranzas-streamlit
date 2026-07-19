from __future__ import annotations

import csv
import os
import random
import re
import sys
import warnings
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = PROJECT_ROOT / "data" / "dataset_cobranzas.xlsx"
LOG_PATH = PROJECT_ROOT / "data" / "update_log.csv"
FACTURAS_SHEET = "Facturas"
FACTURAS_TABLE = "FacturasTable"
KEEP_LAST_MONTHS = int(os.getenv("KEEP_LAST_MONTHS", "12"))
MIN_NEW_RECORDS = int(os.getenv("MIN_NEW_RECORDS", "3"))
MAX_NEW_RECORDS = int(os.getenv("MAX_NEW_RECORDS", "8"))

OBSERVACIONES = [
    "Factura generada automáticamente para portfolio",
    "Movimiento ficticio agregado por rutina diaria",
    "Registro simulado sin datos reales",
    "Caso demo para seguimiento de cobranzas",
]
ACCIONES = {
    "sin_prioridad": "Sin acción requerida",
    "recordatorio": "Enviar recordatorio de pago",
    "responsable": "Contactar responsable administrativo",
    "parcial": "Revisar pago parcial",
    "escalar": "Escalar seguimiento comercial",
    "confirmar": "Confirmar fecha estimada de pago",
}


class DatasetUpdateError(RuntimeError):
    pass


def normalize_name(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def find_facturas_range(workbook: Any) -> tuple[Any, str, tuple[int, int, int, int]]:
    if FACTURAS_SHEET in workbook.sheetnames:
        worksheet = workbook[FACTURAS_SHEET]
        if FACTURAS_TABLE in worksheet.tables:
            table_ref = worksheet.tables[FACTURAS_TABLE].ref
            return worksheet, table_ref, range_boundaries(table_ref)

    for worksheet in workbook.worksheets:
        for table in worksheet.tables.values():
            if normalize_name(table.name) == normalize_name(FACTURAS_TABLE):
                return worksheet, table.ref, range_boundaries(table.ref)

    if FACTURAS_SHEET not in workbook.sheetnames:
        raise DatasetUpdateError("No se encontró la hoja principal 'Facturas'.")

    worksheet = workbook[FACTURAS_SHEET]
    if worksheet.max_row < 2:
        raise DatasetUpdateError("La hoja 'Facturas' no tiene registros para usar como base.")

    return worksheet, f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}", (
        1,
        1,
        worksheet.max_column,
        worksheet.max_row,
    )


def read_range_as_dataframe(worksheet: Any, bounds: tuple[int, int, int, int]) -> pd.DataFrame:
    min_col, min_row, max_col, max_row = bounds
    rows = list(
        worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    )
    if not rows:
        raise DatasetUpdateError("No se pudieron leer filas de la tabla principal.")

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    if not headers or "id_factura" not in headers:
        raise DatasetUpdateError("La tabla principal no contiene la columna requerida 'id_factura'.")

    return pd.DataFrame(rows[1:], columns=headers)


def read_sheet(workbook: Any, sheet_name: str) -> pd.DataFrame:
    if sheet_name not in workbook.sheetnames:
        return pd.DataFrame()
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data = pd.DataFrame(rows[1:], columns=headers)
    return data.loc[:, [column for column in data.columns if column and not column.startswith("Unnamed")]]


def prepare_facturas(df: pd.DataFrame) -> pd.DataFrame:
    prepared = df.copy()
    for column in ["fecha_emision", "fecha_vencimiento", "fecha_pago", "mes_emision", "mes_vencimiento"]:
        if column in prepared.columns:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                prepared[column] = pd.to_datetime(prepared[column], errors="coerce")
    for column in [
        "importe_facturado",
        "importe_cobrado",
        "saldo_pendiente",
        "tipo_cambio",
        "importe_equivalente_ars",
        "saldo_equivalente_ars",
        "dias_atraso",
    ]:
        if column in prepared.columns:
            prepared[column] = pd.to_numeric(prepared[column], errors="coerce").fillna(0)
    return prepared


def next_invoice_ids(existing_ids: pd.Series, quantity: int) -> list[str]:
    numeric_ids = []
    for value in existing_ids.dropna().astype(str):
        match = re.fullmatch(r"FAC-(\d+)", value.strip())
        if match:
            numeric_ids.append(int(match.group(1)))

    if numeric_ids:
        width = max(5, max(len(str(number)) for number in numeric_ids))
        start = max(numeric_ids) + 1
        return [f"FAC-{number:0{width}d}" for number in range(start, start + quantity)]

    today_key = date.today().strftime("%Y%m%d")
    existing = set(existing_ids.dropna().astype(str))
    ids = []
    counter = 1
    while len(ids) < quantity:
        candidate = f"FAC-{today_key}-{counter:03d}"
        if candidate not in existing:
            ids.append(candidate)
        counter += 1
    return ids


def month_start(value: date) -> datetime:
    return datetime(value.year, value.month, 1)


def quarter_label(value: date) -> str:
    return f"{value.year}-T{((value.month - 1) // 3) + 1}"


def clean_records(df: pd.DataFrame, required: list[str]) -> list[dict[str, Any]]:
    if df.empty:
        return []
    available = [column for column in required if column in df.columns]
    records = df.dropna(subset=available[:1]).to_dict("records")
    return [{key: value for key, value in record.items() if pd.notna(value)} for record in records]


def valid_entity_rows(df: pd.DataFrame, id_column: str, prefix: str) -> pd.DataFrame:
    if df.empty or id_column not in df.columns:
        return df
    pattern = rf"^{re.escape(prefix)}-\d+"
    return df[df[id_column].astype(str).str.match(pattern, na=False)].copy()


def payment_days_from_client(client: dict[str, Any]) -> int:
    value = pd.to_numeric(client.get("condicion_pago_dias"), errors="coerce")
    if pd.isna(value) or value <= 0:
        return random.choice([15, 30, 45])
    return int(value)


def historical_amount(facturas: pd.DataFrame, service_id: str, currency: str) -> float:
    scoped = facturas[
        facturas.get("id_servicio", pd.Series(dtype=str)).astype(str).eq(str(service_id))
        & facturas.get("moneda", pd.Series(dtype=str)).astype(str).str.upper().eq(currency)
    ]
    if len(scoped) < 4:
        scoped = facturas[facturas.get("moneda", pd.Series(dtype=str)).astype(str).str.upper().eq(currency)]
    if scoped.empty:
        return 500000 if currency == "ARS" else 650

    low = float(scoped["importe_facturado"].quantile(0.15))
    high = float(scoped["importe_facturado"].quantile(0.85))
    if high <= low:
        high = low * 1.25
    value = random.uniform(low, high)
    if currency == "USD":
        return round(value / 50) * 50
    return round(value / 1000) * 1000


def latest_exchange_rate(workbook: Any) -> float:
    rates = read_sheet(workbook, "Tipo_Cambio")
    if rates.empty or "tipo_cambio_referencia" not in rates.columns:
        return 1300.0
    values = pd.to_numeric(rates["tipo_cambio_referencia"], errors="coerce").dropna()
    if values.empty:
        return 1300.0
    return float(values.iloc[-1]) * random.uniform(1.0, 1.04)


def choose_payment(invoice_amount: float, due_date: date, today: date) -> tuple[float, date | None]:
    roll = random.random()
    if roll < 0.34:
        paid_date = due_date + timedelta(days=random.randint(-4, 18))
        return invoice_amount, paid_date
    if roll < 0.56:
        paid = invoice_amount * random.uniform(0.25, 0.75)
        paid_date = min(today, due_date + timedelta(days=random.randint(-2, 20)))
        return round(paid, 2), paid_date
    return 0.0, None


def classify_invoice(invoice_amount: float, paid: float, balance: float, due_date: date, payment_date: date | None) -> tuple[str, int]:
    today = date.today()
    if paid >= invoice_amount:
        delay = max(((payment_date or today) - due_date).days, 0)
        return "Cobrada", delay

    delay = max((today - due_date).days, 0)
    if paid > 0 and balance > 0:
        return "Parcial", delay
    if due_date >= today:
        return "Pendiente", 0
    if delay >= 45:
        return "En gestión", delay
    return "Vencida", delay


def cobranza_fields(status: str, balance_ars: float, delay: int, reference_high_balance: float) -> tuple[str, str]:
    if status == "Cobrada":
        return "Sin prioridad", ACCIONES["sin_prioridad"]
    if status == "Parcial":
        return "Media", ACCIONES["parcial"]
    if delay >= 45 or balance_ars >= reference_high_balance:
        return "Alta", ACCIONES["escalar"]
    if status == "Vencida":
        return "Alta", ACCIONES["responsable"]
    if balance_ars >= reference_high_balance * 0.65:
        return "Media", ACCIONES["confirmar"]
    return "Baja", ACCIONES["recordatorio"]


def generate_records(facturas: pd.DataFrame, clients_df: pd.DataFrame, services_df: pd.DataFrame, workbook: Any) -> pd.DataFrame:
    clients = clean_records(valid_entity_rows(clients_df, "id_cliente", "CL"), ["id_cliente", "cliente"])
    services = clean_records(valid_entity_rows(services_df, "id_servicio", "SRV"), ["id_servicio", "tipo_servicio"])
    if not clients:
        clients = clean_records(facturas[["id_cliente", "cliente", "tipo_cliente", "responsable"]], ["id_cliente", "cliente"])
    if not services:
        services = clean_records(facturas[["id_servicio", "tipo_servicio"]], ["id_servicio", "tipo_servicio"])
    if not clients or not services:
        raise DatasetUpdateError("No hay clientes o servicios suficientes para generar registros ficticios.")

    quantity = random.randint(MIN_NEW_RECORDS, MAX_NEW_RECORDS)
    invoice_ids = next_invoice_ids(facturas["id_factura"], quantity)
    exchange_rate = latest_exchange_rate(workbook)
    reference_high_balance = float(facturas["saldo_equivalente_ars"].quantile(0.75))
    records = []

    for invoice_id in invoice_ids:
        client = random.choice(clients)
        service = random.choice(services)
        payment_days = payment_days_from_client(client)
        issue_date = date.today() - timedelta(days=random.randint(0, 20))
        due_date = issue_date + timedelta(days=payment_days)
        currency = str(service.get("moneda_principal") or random.choice(["ARS", "USD"])).upper()
        if currency not in {"ARS", "USD"}:
            currency = random.choice(["ARS", "USD"])

        amount = max(historical_amount(facturas, str(service.get("id_servicio", "")), currency), 1)
        paid, payment_date = choose_payment(amount, due_date, date.today())
        balance = max(amount - paid, 0)
        tipo_cambio = round(exchange_rate, 2) if currency == "USD" else 1
        amount_ars = amount * tipo_cambio if currency == "USD" else amount
        balance_ars = balance * tipo_cambio if currency == "USD" else balance
        status, delay = classify_invoice(amount, paid, balance, due_date, payment_date)
        priority, next_action = cobranza_fields(status, balance_ars, delay, reference_high_balance)

        record = {
            "id_factura": invoice_id,
            "fecha_emision": issue_date,
            "fecha_vencimiento": due_date,
            "fecha_pago": payment_date,
            "id_cliente": client.get("id_cliente"),
            "cliente": client.get("cliente"),
            "tipo_cliente": client.get("tipo_cliente", "PyME servicios recurrentes"),
            "id_servicio": service.get("id_servicio"),
            "tipo_servicio": service.get("tipo_servicio"),
            "moneda": currency,
            "importe_facturado": round(amount, 2),
            "importe_cobrado": round(paid, 2),
            "saldo_pendiente": round(balance, 2),
            "tipo_cambio": tipo_cambio,
            "importe_equivalente_ars": round(amount_ars, 2),
            "saldo_equivalente_ars": round(balance_ars, 2),
            "estado": status,
            "dias_atraso": int(delay),
            "prioridad_cobranza": priority,
            "responsable": client.get("responsable", "Equipo administrativo"),
            "proxima_accion": next_action,
            "observaciones": random.choice(OBSERVACIONES),
            "mes_emision": month_start(issue_date),
            "mes_vencimiento": month_start(due_date),
            "trimestre": quarter_label(issue_date),
            "vencida_si_no": "Sí" if delay > 0 and balance > 0 else "No",
            "parcial_si_no": "Sí" if status == "Parcial" else "No",
            "exposicion_usd_si_no": "Sí" if currency == "USD" else "No",
        }
        records.append(record)

    return pd.DataFrame(records)


def prune_records(df: pd.DataFrame) -> pd.DataFrame:
    if KEEP_LAST_MONTHS <= 0 or "fecha_emision" not in df.columns:
        return df
    cutoff = pd.Timestamp(date.today()) - pd.DateOffset(months=KEEP_LAST_MONTHS)
    dates = pd.to_datetime(df["fecha_emision"], errors="coerce")
    return df[dates.ge(cutoff) | dates.isna()].copy()


def dataframe_to_excel_rows(df: pd.DataFrame, columns: list[str]) -> list[list[Any]]:
    rows = []
    for record in df[columns].to_dict("records"):
        row = []
        for value in record.values():
            if pd.isna(value):
                row.append(None)
            elif isinstance(value, pd.Timestamp):
                row.append(value.to_pydatetime())
            else:
                row.append(value)
        rows.append(row)
    return rows


def write_facturas(workbook: Any, worksheet: Any, bounds: tuple[int, int, int, int], df: pd.DataFrame) -> None:
    min_col, min_row, max_col, max_row = bounds
    columns = list(df.columns)
    if len(columns) != (max_col - min_col + 1):
        raise DatasetUpdateError("La cantidad de columnas generada no coincide con la tabla original.")

    if max_row > min_row:
        worksheet.delete_rows(min_row + 1, max_row - min_row)

    for offset, column in enumerate(columns, start=min_col):
        worksheet.cell(row=min_row, column=offset, value=column)

    for row_values in dataframe_to_excel_rows(df, columns):
        worksheet.append(row_values)

    new_max_row = min_row + len(df)
    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
    if FACTURAS_TABLE in worksheet.tables:
        worksheet.tables[FACTURAS_TABLE].ref = new_ref
    worksheet.auto_filter.ref = new_ref
    workbook.save(DATA_PATH)


def append_log(before: int, added: int, after: int, status: str) -> None:
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    exists = LOG_PATH.exists()
    with LOG_PATH.open("a", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "fecha_hora_ejecucion",
                "cantidad_registros_antes",
                "cantidad_registros_agregados",
                "cantidad_registros_despues",
                "archivo_actualizado",
                "estado_ejecucion",
            ],
        )
        if not exists:
            writer.writeheader()
        writer.writerow(
            {
                "fecha_hora_ejecucion": datetime.now().isoformat(timespec="seconds"),
                "cantidad_registros_antes": before,
                "cantidad_registros_agregados": added,
                "cantidad_registros_despues": after,
                "archivo_actualizado": str(DATA_PATH.relative_to(PROJECT_ROOT)),
                "estado_ejecucion": status,
            }
        )


def update_dataset() -> tuple[int, int, int]:
    if not DATA_PATH.exists():
        raise DatasetUpdateError(f"No existe el archivo Excel esperado: {DATA_PATH}")

    workbook = load_workbook(DATA_PATH)
    worksheet, _, bounds = find_facturas_range(workbook)
    facturas = prepare_facturas(read_range_as_dataframe(worksheet, bounds))
    if facturas.empty:
        raise DatasetUpdateError("La tabla de facturas está vacía.")

    before = len(facturas)
    clients_df = read_sheet(workbook, "Clientes")
    services_df = read_sheet(workbook, "Servicios")
    new_records = generate_records(facturas, clients_df, services_df, workbook)
    combined = pd.concat([facturas, new_records], ignore_index=True)
    for column in ["fecha_emision", "fecha_vencimiento", "fecha_pago", "mes_emision", "mes_vencimiento"]:
        if column in combined.columns:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                combined[column] = pd.to_datetime(combined[column], errors="coerce")
    combined = prune_records(combined)
    combined = combined.drop_duplicates(subset=["id_factura"], keep="last")
    combined = combined.sort_values(["fecha_emision", "id_factura"]).reset_index(drop=True)

    write_facturas(workbook, worksheet, bounds, combined)
    after = len(combined)
    append_log(before, len(new_records), after, "ok")
    return before, len(new_records), after


def main() -> int:
    before = after = added = 0
    try:
        before, added, after = update_dataset()
    except Exception as exc:
        append_log(before, added, after, f"error: {exc}")
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(
        f"Dataset actualizado: {before} registros antes, "
        f"{added} agregados, {after} registros después."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
